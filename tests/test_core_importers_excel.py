"""Excel 导入器测试（openpyxl 生成测试文件）。."""

from __future__ import annotations

import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from finaldb.core.exceptions import InvalidDataError, UnsupportedFormatError
from finaldb.core.importers.excel_reader import read_excel, supports_excel


@pytest.fixture()
def xlsx_path(tmp_path: Path) -> Path:
    """生成双 sheet 的测试工作簿。."""
    wb = Workbook()
    ws = wb.active
    ws.title = "人员"
    ws.append(["id", "姓名", "入职日期"])
    ws.append([1, "甲", datetime.date(2024, 1, 15)])
    ws.append([2, "乙", None])
    ws2 = wb.create_sheet("工资")
    ws2.append(["id", "金额"])
    ws2.append([1, 8000.5])
    path = tmp_path / "名单.xlsx"
    wb.save(path)
    return path


def test_supports_excel_suffixes() -> None:
    """扩展名支持判断。."""
    assert supports_excel(".xlsx")
    assert supports_excel(".XLSX")
    assert not supports_excel(".xls")


def test_multi_sheet_iteration(xlsx_path: Path) -> None:
    """多 sheet 逐个产出 TableData，表名 = 文件名_sheet。."""
    tables = list(read_excel(xlsx_path))
    assert len(tables) == 2
    assert tables[0].name == "名单_人员"
    assert tables[0].columns == ("id", "姓名", "入职日期")
    assert tables[1].name == "名单_工资"
    assert tables[1].columns == ("id", "金额")


def test_cell_normalization(xlsx_path: Path) -> None:
    """日期转 ISO 文本，空单元格为 None，数值保留。."""
    tables = list(read_excel(xlsx_path))
    rows = list(tables[0].rows)
    assert rows[0] == (1, "甲", "2024-01-15")
    assert rows[1] == (2, "乙", None)
    assert list(tables[1].rows) == [(1, 8000.5)]


def test_xls_rejected(tmp_path: Path) -> None:
    """旧版 .xls 报不支持。."""
    fake = tmp_path / "old.xls"
    fake.write_bytes(b"fake")
    with pytest.raises(UnsupportedFormatError, match=".xls"):
        list(read_excel(fake))


def test_non_excel_extension_rejected(tmp_path: Path) -> None:
    """非 Excel 扩展名报错。."""
    fake = tmp_path / "d.txt"
    fake.write_text("x", "utf-8")
    with pytest.raises(UnsupportedFormatError, match="非 Excel"):
        list(read_excel(fake))


def test_empty_sheet_raises(tmp_path: Path) -> None:
    """空 sheet 报无数据。."""
    wb = Workbook()
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    with pytest.raises(InvalidDataError, match="无数据"):
        list(read_excel(path))


def test_duplicate_headers_deduped(tmp_path: Path) -> None:
    """重复表头自动去重。."""
    wb = Workbook()
    ws = wb.active
    ws.title = "s"
    ws.append(["id", "id"])
    ws.append([1, 2])
    path = tmp_path / "dup.xlsx"
    wb.save(path)
    td = next(iter(read_excel(path)))
    assert td.columns == ("id", "id_2")
