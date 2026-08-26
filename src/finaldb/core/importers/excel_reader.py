"""Excel 导入器：openpyxl 只读流式解析（x.xlsx/xlsm）。."""

from __future__ import annotations

import datetime as _dt
import decimal
from collections.abc import Iterator
from pathlib import Path

from finaldb.core.exceptions import InvalidDataError, UnsupportedFormatError
from finaldb.core.importers.base import TableData
from finaldb.core.importers.naming import sanitize_identifier

__all__ = ["read_excel", "supports_excel"]

# openpyxl 支持的扩展名
_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
# .xls 为旧二进制格式，openpyxl 不支持
_REJECTED_SUFFIXES = {".xls"}


def supports_excel(suffix: str) -> bool:
    """判断扩展名是否为 openpyxl 支持的 Excel 格式。."""
    return suffix.lower() in _EXCEL_SUFFIXES


def read_excel(path: Path) -> Iterator[TableData]:
    """解析 Excel 工作簿，逐 sheet 产出 TableData（惰性行迭代器）。

    首行为表头；日期时间统一转 ISO 文本；Decimal 转 float。

    :param path: 文件路径（.xlsx/.xlsm）
    :return: 每个 sheet 一个 TableData
    :raises UnsupportedFormatError: 扩展名为 .xls 等不支持格式
    :raises InvalidDataError: 无 sheet 或 sheet 无数据
    """
    suffix = path.suffix.lower()
    if suffix in _REJECTED_SUFFIXES:
        raise UnsupportedFormatError(f"不支持旧版 .xls 格式，请另存为 .xlsx: {path.name}")
    if suffix not in _EXCEL_SUFFIXES:
        raise UnsupportedFormatError(f"非 Excel 文件: {path.name}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - openpyxl 为必装依赖
        raise UnsupportedFormatError("缺少 openpyxl 依赖") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = list(wb.sheetnames)
        if not sheets:
            raise InvalidDataError(f"工作簿无工作表: {path.name}")
        for sheet in sheets:
            ws = wb[sheet]
            table = sanitize_identifier(f"{path.stem}_{sheet}")
            yield _sheet_to_table(table, ws)
    finally:
        wb.close()


def _sheet_to_table(table: str, ws: object) -> TableData:
    """把单个 sheet 转为 TableData（首行表头）。."""
    row_iter = iter(ws.iter_rows(values_only=True))  # pyrefly: ignore [missing-attribute]
    header = next(row_iter, None)
    if header is None:
        raise InvalidDataError(f"工作表无数据: {table}")
    first_row = next(row_iter, None)
    width = len(header)
    if first_row is None and not any(cell is not None and str(cell).strip() for cell in header):
        # 既无表头也无数据：产出空表（0 列）
        raise InvalidDataError(f"工作表无数据: {table}")

    def rows() -> Iterator[tuple[object, ...]]:
        """惰性行迭代器：规范化单元格 + 定宽。."""
        for raw in _chain(first_row, row_iter):
            vals = [_normalize_cell(c) for c in raw[:width]]
            while len(vals) < width:
                vals.append(None)
            yield tuple(vals)

    if any(cell is not None and str(cell).strip() for cell in header):
        columns = _dedupe([str(c) if c is not None else "" for c in header])
    else:
        columns = tuple(f"c{i + 1}" for i in range(width))
    return TableData(name=table, columns=columns, rows=rows())


def _chain(
    first: tuple[object, ...] | None,
    rest: Iterator[tuple[object, ...]],
) -> Iterator[tuple[object, ...]]:
    """首行接回迭代器头部。."""
    if first is not None:
        yield first
    yield from rest


def _dedupe(header: list[str]) -> tuple[str, ...]:
    """表头清洗去重。."""
    from finaldb.core.importers.naming import sanitize_identifier

    cleaned = [sanitize_identifier(h.strip()) for h in header]
    seen: dict[str, int] = {}
    result = []
    for col in cleaned:
        if col in seen:
            seen[col] += 1
            result.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 1
            result.append(col)
    return tuple(result)


def _normalize_cell(value: object) -> object:
    """单元格规范化：日期→ISO 文本，Decimal→float，其余原样。

    零点 datetime（xlsx 中 date 的存储形式）归一为纯日期文本。
    """
    if isinstance(value, _dt.datetime):
        if value.time() == _dt.time(0):
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, (_dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, decimal.Decimal):
        return float(value)
    if isinstance(value, str):
        return value.strip() or None
    return value
